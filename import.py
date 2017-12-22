import csv
import sys
import yaml
from re import sub
from decimal import Decimal
from openpyxl import load_workbook

if len(sys.argv) != 2:
    print "usage: <input-yaml>"
    exit(1)

yaml_file_name = sys.argv[1]


summary = dict()
summary['Branded'] = dict()
summary['Non-Brand'] = dict()
summary['Affiliate'] = dict()
summary['Display'] = dict()
summary['Social'] = dict()
summary['Branded']['clicks'] = 0
summary['Branded']['impressions'] = 0
summary['Branded']['cost'] = 0
summary['Non-Brand']['clicks'] = 0
summary['Non-Brand']['impressions'] = 0
summary['Non-Brand']['cost'] = 0
summary['Display']['clicks'] = 0
summary['Display']['impressions'] = 0
summary['Display']['cost'] = 0
summary['Social']['revenue'] = 0
summary['Social']['visits'] = 0
summary['Social']['orders'] = 0

def comment_stripper (iterator):
    for line in iterator:
        line = unicode(line, 'utf-8-sig')
        if line[:1] == '#':
            continue
        if not line.strip():
            continue
        yield line

def headers_from_sheet(sheet):
    return map(lambda col: col[0].value, sheet.iter_cols(max_row=1))

def find_row_by_week(sheet, week_code):
    index = headers_from_sheet(sheet).index('FW')
    return map(lambda item: item.value, tuple(sheet.columns)[index]).index(week_code) + 1

def parse_google(google_file_name):
    with open(google_file_name, 'rb') as google_file:
        reader = csv.DictReader(google_file.readlines()[2:])

        for row in reader:
            clicks = 0
            if 'Clicks' in row.keys():
                clicks = int(row['Clicks'].replace(',', ''))
            elif 'Interactions' in row.keys():
                clicks = int(row['Interactions'].replace(',', ''))

            impressions = int(row['Impr.'].replace(',', ''))
            cost = Decimal(sub(r'[^\d.]', '', row['Cost']))
            if 'Non-Brand' in row['Campaign']:
                summary['Non-Brand']['clicks'] += clicks
                summary['Non-Brand']['impressions'] += impressions
                summary['Non-Brand']['cost'] += cost
            elif 'Branded' in row['Campaign']:
                summary['Branded']['clicks'] += clicks
                summary['Branded']['impressions'] += impressions
                summary['Branded']['cost'] += cost

def parse_bing(bing_file_name):
    with open(bing_file_name) as bing_file:
        reader = csv.DictReader(bing_file.readlines()[3:])
        for row in reader:
            clicks = int(row['Clicks'].replace(',', ''))
            impressions = int(row['Impr.'].replace(',', ''))
            cost = Decimal(sub(r'[^\d.]', '', row['Spend']))
            if 'Non-Brand' in row['Campaign']:
                summary['Non-Brand']['clicks'] += clicks
                summary['Non-Brand']['impressions'] += impressions
                summary['Non-Brand']['cost'] += cost
            elif 'Branded' in row['Campaign']:
                summary['Branded']['clicks'] += clicks
                summary['Branded']['impressions'] += impressions
                summary['Branded']['cost'] += cost

def parse_omniture(omniture_file_name):
    with open(omniture_file_name) as omniture_file:
        reader = csv.DictReader(comment_stripper(omniture_file))
        for row in reader:
            sheet_name = row['Last Touch Channel']
            revenue = int(row['Revenue'])
            visits = int(row['Visits'])
            orders = int(row['Orders'])
            if sheet_name == 'Affiliate' or sheet_name == 'Affliate':
                summary['Affiliate']['revenue'] = revenue
                summary['Affiliate']['visits'] = visits
                summary['Affiliate']['orders'] = orders
            elif sheet_name == 'Display':
                summary['Display']['revenue'] = revenue
                summary['Display']['visits'] = visits
                summary['Display']['orders'] = orders
            elif sheet_name == 'Paid Search Branded':
                summary['Branded']['revenue'] = revenue
                summary['Branded']['visits'] = visits
                summary['Branded']['orders'] = orders
            elif sheet_name == 'Paid Search Unbranded':
                summary['Non-Brand']['revenue'] = revenue
                summary['Non-Brand']['visits'] = visits
                summary['Non-Brand']['orders'] = orders
            elif 'Paid Social' in sheet_name:
                summary['Social']['revenue'] += revenue
                summary['Social']['visits'] += visits
                summary['Social']['orders'] += orders

def parse_display(display_file_name):
    with open(display_file_name, 'rb') as display_file_name:
        reader = csv.DictReader(display_file_name.readlines()[2:])

        for row in reader:
            clicks = 0
            if 'Clicks' in row.keys():
                clicks = int(row['Clicks'].replace(',', ''))
            elif 'Interactions' in row.keys():
                clicks = int(row['Interactions'].replace(',', ''))

            impressions = int(row['Impr.'].replace(',', ''))
            cost = Decimal(sub(r'[^\d.]', '', row['Cost']))
            summary['Display']['clicks'] += clicks
            summary['Display']['impressions'] += impressions
            summary['Display']['cost'] += cost
    

report_file_name = ''
Week = ''
search_sheet_name = 'Search'
affiliate_sheet_name = 'Affiliate'
display_sheet_name = 'Display - Promos'
social_sheet_name = ''

with open(yaml_file_name) as stream:
    reader = yaml.load(stream)
    report_file_name = reader['source']
    Week = reader['week']
    
    if 'google' in reader:
        parse_google(reader['google'])

    if 'bing' in reader:
        parse_bing(reader['bing'])

    if 'omniture' in reader:    
        parse_omniture(reader['omniture'])

    if 'display' in reader:
        parse_display(reader['display'])

    if 'banner' in reader:
        search_sheet_name = reader['banner'] + ' SEM'
        affiliate_sheet_name = reader['banner'] + ' AFFILIATES'
        display_sheet_name = reader['banner'] + ' DISPLAY'
        social_sheet_name = reader['banner'] + ' SOCIAL'

print summary

wb = load_workbook(report_file_name)
search_sheet = wb[search_sheet_name]

headers = headers_from_sheet(search_sheet)

base_row = find_row_by_week(search_sheet, Week)
branded_row = base_row + 1
non_brand_row = base_row + 2
cost_col = headers.index('Spend') + 1
clicks_col = headers.index('Clicks') + 1
try:
    impr_col = headers.index('Impressions') + 1
except:
    impr_col = headers.index('Imp') + 1

revenue_col = headers.index('Revenue') + 1
visits_col = headers.index('Visits') + 1
orders_col = headers.index('Orders') + 1

search_sheet.cell(row=branded_row, column=clicks_col, value = summary['Branded']['clicks'])
search_sheet.cell(row=branded_row, column=impr_col, value = summary['Branded']['impressions'])
search_sheet.cell(row=branded_row, column=cost_col, value = summary['Branded']['cost'])
search_sheet.cell(row=branded_row, column=revenue_col, value = summary['Branded']['revenue'])
search_sheet.cell(row=branded_row, column=visits_col, value = summary['Branded']['visits'])
search_sheet.cell(row=branded_row, column=orders_col, value = summary['Branded']['orders'])
search_sheet.cell(row=non_brand_row, column=clicks_col, value = summary['Non-Brand']['clicks'])
search_sheet.cell(row=non_brand_row, column=impr_col, value = summary['Non-Brand']['impressions'])
search_sheet.cell(row=non_brand_row, column=cost_col, value = summary['Non-Brand']['cost'])
search_sheet.cell(row=non_brand_row, column=revenue_col, value = summary['Non-Brand']['revenue'])
search_sheet.cell(row=non_brand_row, column=visits_col, value = summary['Non-Brand']['visits'])
search_sheet.cell(row=non_brand_row, column=orders_col, value = summary['Non-Brand']['orders'])

affiliate_sheet = wb[affiliate_sheet_name]
headers = headers_from_sheet(affiliate_sheet)

base_row = find_row_by_week(affiliate_sheet, Week)
revenue_col = headers.index('Revenue') + 1
visits_col = headers.index('Visits') + 1
orders_col = headers.index('Orders') + 1

affiliate_sheet.cell(row=base_row, column=revenue_col, value = summary['Affiliate']['revenue'])
affiliate_sheet.cell(row=base_row, column=visits_col, value = summary['Affiliate']['visits'])
affiliate_sheet.cell(row=base_row, column=orders_col, value = summary['Affiliate']['orders'])

display_sheet = wb[display_sheet_name]
headers = headers_from_sheet(display_sheet)

if (display_sheet_name == 'Display - Promos'):
    base_row = find_row_by_week(display_sheet, Week) + 2
else:
    base_row = find_row_by_week(display_sheet, Week)

revenue_col = headers.index('Revenue') + 1
visits_col = headers.index('Visits') + 1
orders_col = headers.index('Orders') + 1

display_sheet.cell(row=base_row, column=revenue_col, value = summary['Display']['revenue'])
display_sheet.cell(row=base_row, column=visits_col, value = summary['Display']['visits'])
display_sheet.cell(row=base_row, column=orders_col, value = summary['Display']['orders'])

social_sheet = wb[social_sheet_name]
headers = headers_from_sheet(social_sheet)

if (social_sheet_name != ''):
    base_row = find_row_by_week(social_sheet, Week)
    revenue_col = headers.index('Revenue') + 1
    visits_col = headers.index('Visits') + 1
    orders_col = headers.index('Orders') + 1

    social_sheet.cell(row=base_row, column=revenue_col, value = summary['Social']['revenue'])
    social_sheet.cell(row=base_row, column=visits_col, value = summary['Social']['visits'])
    social_sheet.cell(row=base_row, column=orders_col, value = summary['Social']['orders'])


wb.save('output.xlsx')
